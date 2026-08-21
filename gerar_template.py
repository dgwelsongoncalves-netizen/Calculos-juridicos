from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

f_bold = Font(bold=True)
borda = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
fill_cinza = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

# --- 1. ABA PARAMETROS ---
ws_param = wb.active
ws_param.title = "Parametros"

parametros = [
    ("Processo", "5000000-00.2026.8.13.0024"),
    ("Atuação", "RÉU"),  # AUTOR ou RÉU
    ("Data do Trânsito", "30/08/2023"),
    ("Data da Sentença", "18/07/2023"),
    ("Termo Inicial Juros", "Desembolso"),
    ("Data da Citação", "19/05/2022"),
    ("Data do Evento", "05/09/2016"),
    ("Justiça Gratuita", "Não"),
    ("Pagamento Voluntário 15d", "Sim"),
    ("Fazenda Pública", "Não"),
    ("Honorários Sucumbência", 10),
    ("Honorários Fixos (R$)", ""),
    ("Base Honorários", "CONDENAÇÃO"),
    ("Valor Causa Original", ""),
    ("Data Propositura", ""),
    ("Proporção Honorários (%)", "100%"),
    ("Proporção Custas (%)", "100%")
]

for r_idx, (chave, valor) in enumerate(parametros, 1):
    c1 = ws_param.cell(row=r_idx, column=1, value=chave)
    c1.font = f_bold; c1.fill = fill_cinza; c1.border = borda
    
    c2 = ws_param.cell(row=r_idx, column=2, value=valor)
    c2.border = borda

ws_param.column_dimensions['A'].width = 25
ws_param.column_dimensions['B'].width = 30

# --- 2. ABA DANOS ---
ws_danos = wb.create_sheet('Danos')
headers_danos = ['ID / Folha', 'Descrição', 'Data Desembolso', 'Valor Histórico', 'Regra', 'Data Juros', 'Valor Pedido Inicial', 'Data do Pedido']
for col_idx, h in enumerate(headers_danos, 1):
    cell = ws_danos.cell(row=1, column=col_idx, value=h)
    cell.font = f_bold; cell.fill = fill_cinza; cell.border = borda

for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
    ws_danos.column_dimensions[col_letter].width = 20

# --- 3. ABA CUSTAS ---
ws_custas = wb.create_sheet('Custas')
headers_custas = ['ID / Folha', 'Descrição', 'Data Desembolso', 'Valor Histórico']
for col_idx, h in enumerate(headers_custas, 1):
    cell = ws_custas.cell(row=1, column=col_idx, value=h)
    cell.font = f_bold; cell.fill = fill_cinza; cell.border = borda

for col_letter in ['A', 'B', 'C', 'D']:
    ws_custas.column_dimensions[col_letter].width = 22

# --- 4. ABA DEDUCOES ---
ws_deducoes = wb.create_sheet('Deducoes')
headers_deducoes = ['ID / Folha', 'Data bloqueio/deposito', 'Valor']
for col_idx, h in enumerate(headers_deducoes, 1):
    cell = ws_deducoes.cell(row=1, column=col_idx, value=h)
    cell.font = f_bold; cell.fill = fill_cinza; cell.border = borda

for col_letter in ['A', 'B', 'C']:
    ws_deducoes.column_dimensions[col_letter].width = 25

# Salva o arquivo oficial
wb.save('template_nash.xlsx')
print("Template oficial corrigido e salvo como 'template_nash.xlsx' com sucesso!")