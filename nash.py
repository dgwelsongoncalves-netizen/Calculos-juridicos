import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import tkinter as tk
from tkinter import filedialog, messagebox
import threading
import sys
import os
import logging
import traceback
import platform
import subprocess

__version__ = "2.8.4" # Add: Parâmetro de Atuação (Autor/Réu) e Desacoplamento Total de Marcos (Correção vs Juros por Verba)

# --- 1. CONFIGURAÇÕES BASE ---
if getattr(sys, 'frozen', False):
    PASTA_APP = Path(sys.executable).parent
    PASTA_TEMP = Path(sys._MEIPASS) if hasattr(sys, '_MEIPASS') else PASTA_APP
else:
    PASTA_APP = Path(__file__).parent
    PASTA_TEMP = PASTA_APP

PASTA_TABELAS = PASTA_APP / 'Tabelas_Oficiais'
ARQUIVO_TJMG = PASTA_TABELAS / 'tabela_tjmg.xlsx'

ARQUIVO_LOG = PASTA_APP / "nash_system.log"
logging.basicConfig(filename=ARQUIVO_LOG, level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s", datefmt="%d/%m/%Y %H:%M:%S", encoding="utf-8")
logging.info(f"--- Nash System v{__version__} Iniciado ---")

# --- 2. CARREGAMENTO DE DADOS ---
def carregar_tjmg():
    try:
        if not ARQUIVO_TJMG.exists(): raise FileNotFoundError("O arquivo 'tabela_tjmg.xlsx' não foi encontrado na pasta 'Tabelas_Oficiais'.")
        df = pd.read_excel(ARQUIVO_TJMG, sheet_name='Plan1', skiprows=8, names=['ANO', 'MÊS', 'ÍNDICE']).dropna(subset=['MÊS', 'ÍNDICE'])
        meses = {'Janeiro': '01', 'Fevereiro': '02', 'Março': '03', 'Abril': '04', 'Maio': '05', 'Junho': '06', 'Julho': '07', 'Agosto': '08', 'Setembro': '09', 'Outubro': '10', 'Novembro': '11', 'Dezembro': '12'}
        df['MÊS_NUM'] = df['MÊS'].str.strip().map(meses)
        df = df.dropna(subset=['MÊS_NUM'])
        df['DATA_REF'] = pd.to_datetime(df['ANO'].astype(int).astype(str) + '-' + df['MÊS_NUM'] + '-01')
        return df[['DATA_REF', 'ÍNDICE']].set_index('DATA_REF')
    except Exception as e: raise Exception(f"Falha ao carregar tabela do TJMG:\n{e}")

def carregar_taxas_bcb(data_minima):
    try:
        from bcb import sgs
        str_data_inicio = '2010-01-01' if pd.isna(data_minima) else f"{data_minima.year}-{data_minima.month:02d}-01"
        df_selic = sgs.get({'SELIC': 4390}, start=str_data_inicio).apply(pd.to_numeric, errors='coerce').fillna(0) / 100.0
        df_selic.index = df_selic.index.to_period('M').to_timestamp()
        df_ipca = sgs.get({'IPCA': 433}, start=str_data_inicio).apply(pd.to_numeric, errors='coerce').fillna(0) / 100.0
        df_ipca.index = df_ipca.index.to_period('M').to_timestamp()
        inicio_tl = max(pd.to_datetime('2024-08-01'), pd.to_datetime(str_data_inicio))
        df_tl = sgs.get({'TAXA_LEGAL': 29543}, start=inicio_tl.strftime('%Y-%m-%d')).apply(pd.to_numeric, errors='coerce').fillna(0) / 100.0
        df_tl.index = df_tl.index.to_period('M').to_timestamp()
        df_ipca_e = sgs.get({'IPCA_E': 10764}, start=str_data_inicio).apply(pd.to_numeric, errors='coerce').fillna(0) / 100.0
        df_ipca_e.index = df_ipca_e.index.to_period('M').to_timestamp()
        df_poup = sgs.get({'POUPANCA': 195}, start=str_data_inicio).resample('MS').first().apply(pd.to_numeric, errors='coerce').fillna(0) / 100.0
        df_poup.index = df_poup.index.to_period('M').to_timestamp()
        return {'SELIC': df_selic, 'IPCA': df_ipca, 'TAXA_LEGAL': df_tl, 'IPCA_E': df_ipca_e, 'POUPANCA': df_poup}
    except Exception as e: raise Exception(f"Sem conexão API do Banco Central.\nDetalhe: {e}")
    
def obter_fator_tjmg(df_tjmg, data_inicio, data_fim):
    if df_tjmg is None: return 1.0
    d_ini = pd.to_datetime(f"{data_inicio.year}-{data_inicio.month:02d}-01")
    d_fim = pd.to_datetime(f"{data_fim.year}-{data_fim.month:02d}-01")
    if d_fim > df_tjmg.index[-1] + pd.DateOffset(months=2): raise ValueError(f"A Tabela TJMG está defasada!")
    try: idx_ini = float(df_tjmg.loc[d_ini, 'ÍNDICE'])
    except KeyError: idx_ini = float(df_tjmg['ÍNDICE'].iloc[-1])
    try: idx_fim = float(df_tjmg.loc[d_fim, 'ÍNDICE'])
    except KeyError: idx_fim = float(df_tjmg['ÍNDICE'].iloc[-1])
    return idx_ini / idx_fim if idx_fim != 0 else 1.0

# --- 3. MOTORES MATEMÁTICOS ---
def calc_tjmg_juros(df_tjmg, data_cm, data_juros, data_calculo):
    fator_cm = obter_fator_tjmg(df_tjmg, data_cm, data_calculo)
    fator_juros = max(0, (data_calculo.year - data_juros.year) * 12 + (data_calculo.month - data_juros.month)) * 0.01 if pd.notna(data_juros) and data_juros <= data_calculo else 0.0
    return fator_cm, fator_juros

def calc_selic_pura(df_bcb, data_cm, data_juros, data_calculo):
    if df_bcb is None: return 1.0, 0.0
    data_cm_mes, data_calc_mes = pd.to_datetime(f"{data_cm.year}-{data_cm.month:02d}-01"), pd.to_datetime(f"{data_calculo.year}-{data_calculo.month:02d}-01")
    if pd.isna(data_juros) or data_juros > data_calculo: return (1 + df_bcb['IPCA'].loc[(df_bcb['IPCA'].index >= data_cm_mes) & (df_bcb['IPCA'].index <= data_calc_mes), 'IPCA']).prod(), 0.0
    data_juros_mes = pd.to_datetime(f"{data_juros.year}-{data_juros.month:02d}-01")
    fator_ipca = (1 + df_bcb['IPCA'].loc[(df_bcb['IPCA'].index >= data_cm_mes) & (df_bcb['IPCA'].index < data_juros_mes), 'IPCA']).prod() if data_cm_mes < data_juros_mes else 1.0
    fator_selic = (1 + df_bcb['SELIC'].loc[(df_bcb['SELIC'].index >= max(data_cm_mes, data_juros_mes)) & (df_bcb['SELIC'].index <= data_calc_mes), 'SELIC']).prod()
    return fator_ipca, (fator_selic - 1.0)

def calc_tjmg_leinova(df_tjmg, df_bcb, data_cm, data_juros, data_calculo):
    data_corte, corte_mes = pd.to_datetime("2024-08-30"), pd.to_datetime("2024-08-01")
    if data_cm >= data_corte: return calc_leinova_pura(df_bcb, data_cm, data_juros, data_calculo)
    f_cm_1 = obter_fator_tjmg(df_tjmg, data_cm, data_corte)
    jur_1 = max(0, (data_corte.year - data_juros.year) * 12 + (data_corte.month - data_juros.month)) * 0.01 if pd.notna(data_juros) and data_juros < data_corte else 0.0
    data_calc_mes = pd.to_datetime(f"{data_calculo.year}-{data_calculo.month:02d}-01")
    f_cm_2, jur_2 = 1.0, 0.0
    if data_calc_mes > corte_mes and df_bcb is not None:
        f_cm_2 = (1 + df_bcb['IPCA'].loc[(df_bcb['IPCA'].index > corte_mes) & (df_bcb['IPCA'].index <= data_calc_mes), 'IPCA']).prod()
        if pd.notna(data_juros) and data_juros <= data_calculo:
            jur_2 = df_bcb['TAXA_LEGAL'].loc[(df_bcb['TAXA_LEGAL'].index > max(corte_mes, pd.to_datetime(f"{data_juros.year}-{data_juros.month:02d}-01"))) & (df_bcb['TAXA_LEGAL'].index <= data_calc_mes), 'TAXA_LEGAL'].sum()
    return f_cm_1 * f_cm_2, jur_1 + jur_2

def calc_fazenda_publica(df_bcb, data_cm, data_juros, data_calculo):
    if df_bcb is None: return 1.0, 0.0
    corte, mes_corte = pd.to_datetime("2021-11-30"), pd.to_datetime("2021-12-01")
    f_cm_1, jur_1 = 1.0, 0.0
    if data_cm <= corte:
        fim_f1 = min(pd.to_datetime(f"{data_calculo.year}-{data_calculo.month:02d}-01"), mes_corte)
        f_cm_1 = (1 + df_bcb['IPCA_E'].loc[(df_bcb['IPCA_E'].index >= pd.to_datetime(f"{data_cm.year}-{data_cm.month:02d}-01")) & (df_bcb['IPCA_E'].index < fim_f1), 'IPCA_E']).prod()
        if pd.notna(data_juros) and data_juros <= corte:
            jur_1 = df_bcb['POUPANCA'].loc[(df_bcb['POUPANCA'].index >= pd.to_datetime(f"{data_juros.year}-{data_juros.month:02d}-01")) & (df_bcb['POUPANCA'].index < fim_f1), 'POUPANCA'].sum()
    data_calc_mes = pd.to_datetime(f"{data_calculo.year}-{data_calculo.month:02d}-01")
    f_cm_2 = 1.0
    if data_calc_mes >= mes_corte:
        ini_f2 = max(mes_corte, pd.to_datetime(f"{data_cm.year}-{data_cm.month:02d}-01"))
        if pd.isna(data_juros) or data_juros <= mes_corte:
            f_cm_2 = (1 + df_bcb['SELIC'].loc[(df_bcb['SELIC'].index >= ini_f2) & (df_bcb['SELIC'].index <= data_calc_mes), 'SELIC']).prod()
        else:
            d_jur = pd.to_datetime(f"{data_juros.year}-{data_juros.month:02d}-01")
            f_cm_2 = (1 + df_bcb['IPCA_E'].loc[(df_bcb['IPCA_E'].index >= ini_f2) & (df_bcb['IPCA_E'].index < d_jur), 'IPCA_E']).prod() if ini_f2 < d_jur else 1.0
            f_cm_2 *= (1 + df_bcb['SELIC'].loc[(df_bcb['SELIC'].index >= max(ini_f2, d_jur)) & (df_bcb['SELIC'].index <= data_calc_mes), 'SELIC']).prod()
    return f_cm_1 * f_cm_2, jur_1

def calc_leinova_pura(df_bcb, data_cm, data_juros, data_calculo):
    if df_bcb is None: return 1.0, 0.0
    d_cm_m, d_calc_m = pd.to_datetime(f"{data_cm.year}-{data_cm.month:02d}-01"), pd.to_datetime(f"{data_calculo.year}-{data_calculo.month:02d}-01")
    f_ipca = (1 + df_bcb['IPCA'].loc[(df_bcb['IPCA'].index >= d_cm_m) & (df_bcb['IPCA'].index <= d_calc_m), 'IPCA']).prod()
    juros = df_bcb['TAXA_LEGAL'].loc[(df_bcb['TAXA_LEGAL'].index >= pd.to_datetime(f"{data_juros.year}-{data_juros.month:02d}-01")) & (df_bcb['TAXA_LEGAL'].index <= d_calc_m), 'TAXA_LEGAL'].sum() if pd.notna(data_juros) and data_juros <= data_calculo else 0.0
    return f_ipca, juros

# --- 4. EXPORTAÇÃO PARA PDF ---
def converter_para_pdf(caminho_xlsx):
    caminho_xlsx = Path(caminho_xlsx)
    pasta_saida = caminho_xlsx.parent
    
    if platform.system() == "Windows":
        caminhos_lo = [r"C:\Program Files\LibreOffice\program\soffice.exe", r"C:\Program Files (x86)\LibreOffice\program\soffice.exe"]
        lo_path = None
        for p in caminhos_lo:
            if os.path.exists(p):
                lo_path = p
                break
        if not lo_path:
            logging.warning("LibreOffice não encontrado no Windows. PDF não será gerado automaticamente.")
            return False
        comando = [lo_path, "--headless", "--convert-to", "pdf", "--outdir", str(pasta_saida), str(caminho_xlsx)]
    else:
        comando = ["libreoffice", "--headless", "--convert-to", "pdf", "--outdir", str(pasta_saida), str(caminho_xlsx)]
        
    try:
        subprocess.run(comando, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        logging.info(f"PDF gerado com sucesso: {caminho_xlsx.with_suffix('.pdf').name}")
        return True
    except Exception as e:
        logging.error(f"Erro ao converter {caminho_xlsx.name} para PDF: {e}")
        return False

# --- 5. PROCESSAMENTO CENTRAL ---
def executar_nash(caminho_entrada, arquivo_saida):
    tabela_tjmg = carregar_tjmg()
    xls = pd.ExcelFile(caminho_entrada)
    if not all(aba in xls.sheet_names for aba in ['Parametros', 'Danos', 'Custas']): raise Exception("Arquivo inválido. Faltam abas obrigatórias.")

    df_param = pd.read_excel(xls, sheet_name='Parametros', header=None, index_col=0)
    def get_param(nome, default=None, is_date=False):
        try:
            val = df_param.loc[nome, 1]
            if pd.isna(val) or str(val).strip() == '': return default
            return pd.to_datetime(val, dayfirst=True) if is_date else val
        except KeyError: return default

    def parse_percent(val):
        try:
            num = float(str(val).replace('%', '').strip())
            return num / 100.0 if num > 1 else num
        except: return 1.0

    processo = str(get_param('Processo', 'N/A'))
    atuacao = str(get_param('Atuação', 'RÉU')).strip().upper() # NOVO: AUTOR ou RÉU
    data_transito_c, teve_transito = get_param('Data do Trânsito', is_date=True), pd.notna(get_param('Data do Trânsito', is_date=True))
    data_sentenca = get_param('Data da Sentença', is_date=True)
    jg = str(get_param('Justiça Gratuita', '')).strip().upper() == 'SIM'
    houve_inadimplemento = str(get_param('Pagamento Voluntário 15d', '')).strip().upper() in ['NÃO', 'NAO']
    is_fazenda = str(get_param('Fazenda Pública', 'NÃO')).strip().upper() == 'SIM'
    if is_fazenda: houve_inadimplemento = False
        
    hon_perc, hon_fixo = float(get_param('Honorários Sucumbência', 0.0)) / 100, float(get_param('Honorários Fixos (R$)', 0.0))
    termo_juros_raw = str(get_param('Termo Inicial Juros', 'DESEMBOLSO')).strip().upper()
    data_citacao, data_evento = get_param('Data da Citação', is_date=True), get_param('Data do Evento', is_date=True)

    base_hon = str(get_param('Base Honorários', 'CONDENAÇÃO')).strip().upper()
    valor_causa = float(get_param('Valor Causa Original', 0.0))
    data_propositura = get_param('Data Propositura', is_date=True)
    prop_hon, prop_custas = parse_percent(get_param('Proporção Honorários (%)', 100)), parse_percent(get_param('Proporção Custas (%)', 100))

    df_danos = pd.read_excel(xls, sheet_name='Danos').dropna(subset=['Descrição'], how='any')
    df_danos['Valor Histórico'] = pd.to_numeric(df_danos['Valor Histórico'], errors='coerce').fillna(0.0)
    df_danos['Valor Pedido Inicial'] = pd.to_numeric(df_danos.get('Valor Pedido Inicial', 0), errors='coerce').fillna(0.0)
    df_danos['Data Desembolso'] = pd.to_datetime(df_danos.get('Data Desembolso', pd.NaT), format='mixed', dayfirst=True, errors='coerce')
    df_danos['Data do Pedido'] = pd.to_datetime(df_danos.get('Data do Pedido', pd.NaT), format='mixed', dayfirst=True, errors='coerce')
    
    df_danos.loc[(df_danos['Valor Histórico'] == 0) & (df_danos['Data Desembolso'].isna()), 'Data Desembolso'] = df_danos['Data do Pedido']
    
    df_custas = pd.read_excel(xls, sheet_name='Custas').dropna(subset=['Data Desembolso', 'Valor Histórico'], how='any')
    df_custas['Data Desembolso'] = pd.to_datetime(df_custas['Data Desembolso'], format='mixed', dayfirst=True, errors='coerce')
    
    datas = pd.concat([df_danos['Data Desembolso'], df_custas['Data Desembolso'], pd.Series([data_citacao, data_evento, data_propositura])]).dropna()
    df_bcb = carregar_taxas_bcb(datas.min() if not datas.empty else pd.NaT)
    data_calculo = pd.Timestamp.today()
    
    for idx, row in df_danos.iterrows():
        regra_txt = str(row.get('Regra', '')).strip().upper()
        if is_fazenda: df_danos.at[idx, 'Desc_Regra'] = "Faz. Pub. (EC 113)"
        elif regra_txt == 'R1': df_danos.at[idx, 'Desc_Regra'] = "TJMG + Juros 1%"
        elif regra_txt == 'R2': df_danos.at[idx, 'Desc_Regra'] = "Taxa Selic"
        elif regra_txt == 'R3': df_danos.at[idx, 'Desc_Regra'] = "TJMG + 1% até 08/24; após, Selic"
        elif regra_txt == 'R4': df_danos.at[idx, 'Desc_Regra'] = "TJMG + 1% até 08/24; após, Lei 14.905"
        elif regra_txt == 'R5': df_danos.at[idx, 'Desc_Regra'] = "Selic até 08/24; após, Lei 14.905"
        elif regra_txt == 'R6': df_danos.at[idx, 'Desc_Regra'] = "Lei 14.905/24"
        else: df_danos.at[idx, 'Desc_Regra'] = regra_txt if regra_txt else "Selic"
        
    for idx, row in df_custas.iterrows(): df_custas.at[idx, 'Desc_Regra'] = "IPCA+TL (Trânsito)" if teve_transito else "IPCA (S/ Juros)"

    try:
        df_deducoes = pd.read_excel(xls, sheet_name='Deducoes').dropna(subset=['Data bloqueio/deposito', 'Valor'], how='any')
        df_deducoes['Data bloqueio/deposito'] = pd.to_datetime(df_deducoes['Data bloqueio/deposito'], format='mixed', dayfirst=True, errors='coerce')
        df_deducoes = df_deducoes.sort_values('Data bloqueio/deposito')
        tem_deducao = not df_deducoes.empty
    except: tem_deducao = False; df_deducoes = pd.DataFrame()

    total_final_processo = 0.0
    
    for idx, row in df_danos.iterrows():
        data_cm = row['Data Desembolso']
        if pd.isna(data_cm): continue
        valor = float(row['Valor Histórico'])
        regra = str(row.get('Regra', '')).strip().upper()
        
        # SUPORTE A MARCOS INDIVIDUAIS (CORREÇÃO VS JUROS SEPARADOS NA LINHA)
        data_juros_base = data_citacao if 'CITA' in termo_juros_raw else data_evento if 'EVENTO' in termo_juros_raw else data_cm
        data_juros = pd.to_datetime(row['Data Juros'], dayfirst=True, errors='coerce') if 'Data Juros' in row and pd.notna(row['Data Juros']) else data_juros_base
            
        if is_fazenda: f_cm, f_jur = calc_fazenda_publica(df_bcb, data_cm, data_juros, data_calculo)
        elif regra == 'R1': f_cm, f_jur = calc_tjmg_juros(tabela_tjmg, data_cm, data_juros, data_calculo)
        elif regra == 'R4': f_cm, f_jur = calc_tjmg_leinova(tabela_tjmg, df_bcb, data_cm, data_juros, data_calculo)
        elif regra == 'R6': f_cm, f_jur = calc_leinova_pura(df_bcb, data_cm, data_juros, data_calculo)
        else: f_cm, f_jur = calc_selic_pura(df_bcb, data_cm, data_juros, data_calculo)

        val_princ = valor * f_cm
        val_jur = val_princ * f_jur
        df_danos.at[idx, 'Valor Atualizado'] = val_princ + val_jur
        df_danos.at[idx, 'Fator CM'] = f_cm
        df_danos.at[idx, 'Fator Juros'] = f_jur
        
        # CÁLCULO DE ÊXITO APENAS SE ESTIVERMOS PELA DEFESA/RÉU
        v_pedido = float(row['Valor Pedido Inicial'])
        if 'AUTOR' not in atuacao and v_pedido > 0 and pd.notna(row['Data do Pedido']):
            data_ped = row['Data do Pedido']
            if is_fazenda: f_cm_ped, f_jur_ped = calc_fazenda_publica(df_bcb, data_ped, data_juros, data_calculo)
            elif regra == 'R1': f_cm_ped, f_jur_ped = calc_tjmg_juros(tabela_tjmg, data_ped, data_juros, data_calculo)
            elif regra == 'R4': f_cm_ped, f_jur_ped = calc_tjmg_leinova(tabela_tjmg, df_bcb, data_ped, data_juros, data_calculo)
            elif regra == 'R6': f_cm_ped, f_jur_ped = calc_leinova_pura(df_bcb, data_ped, data_juros, data_calculo)
            else: f_cm_ped, f_jur_ped = calc_selic_pura(df_bcb, data_ped, data_juros, data_calculo)
            
            risco_princ = v_pedido * f_cm_ped
            risco_atualizado_total = risco_princ + (risco_princ * f_jur_ped)
            proveito = max(0, risco_atualizado_total - (val_princ + val_jur))
            
            df_danos.at[idx, 'Risco Atual'] = risco_atualizado_total
            df_danos.at[idx, 'Proveito'] = proveito

    if not tem_deducao:
        total_princ_danos = sum([r['Valor Atualizado'] / (1 + r['Fator Juros']) for _, r in df_danos.iterrows() if r['Valor Histórico']>0])
        total_juros_danos = df_danos['Valor Atualizado'].sum() - total_princ_danos
        
        total_princ_custas = total_juros_custas = 0.0
        for idx, row in df_custas.iterrows():
            if jg: df_custas.at[idx, 'Exigível'] = 0.0; continue
            f_cm, f_jur = calc_fazenda_publica(df_bcb, row['Data Desembolso'], data_transito_c if teve_transito else pd.NaT, data_calculo) if is_fazenda else calc_leinova_pura(df_bcb, row['Data Desembolso'], data_transito_c if teve_transito else pd.NaT, data_calculo) 
            val_princ = (row['Valor Histórico'] * f_cm) * prop_custas
            val_jur = val_princ * f_jur
            df_custas.at[idx, 'Exigível'] = val_princ + val_jur
            df_custas.at[idx, 'Fator CM'] = f_cm; df_custas.at[idx, 'Fator Juros'] = f_jur
            total_princ_custas += val_princ; total_juros_custas += val_jur

        subtotal_princ = total_princ_danos + total_princ_custas
        subtotal_juros = total_juros_danos + total_juros_custas
        
        hon_calc_princ, hon_calc_juros = 0.0, 0.0
        if base_hon == 'VALOR DA CAUSA' and valor_causa > 0 and pd.notna(data_propositura):
            f_cm_hon, _ = calc_fazenda_publica(df_bcb, data_propositura, pd.NaT, data_calculo) if is_fazenda else calc_tjmg_leinova(tabela_tjmg, df_bcb, data_propositura, pd.NaT, data_calculo)
            hon_calc_princ = (valor_causa * f_cm_hon) * hon_perc
        elif hon_fixo > 0 and pd.notna(data_sentenca):
            f_cm_hon, f_jur_hon = calc_fazenda_publica(df_bcb, data_sentenca, data_transito_c if teve_transito else pd.NaT, data_calculo) if is_fazenda else calc_leinova_pura(df_bcb, data_sentenca, data_transito_c if teve_transito else pd.NaT, data_calculo)
            hon_calc_princ = hon_fixo * f_cm_hon; hon_calc_juros = hon_calc_princ * f_jur_hon
        elif hon_fixo > 0: hon_calc_princ = hon_fixo
        else: hon_calc_princ = (subtotal_princ + subtotal_juros) * hon_perc
            
        hon_calc_princ *= prop_hon; hon_calc_juros *= prop_hon
        hon_exigivel_princ = 0.0 if jg else hon_calc_princ
        hon_exigivel_juros = 0.0 if jg else hon_calc_juros

        base_multa = subtotal_princ + subtotal_juros + hon_exigivel_princ + hon_exigivel_juros
        valor_multa = (base_multa * 0.10) if houve_inadimplemento else 0.0
        hon_523 = (base_multa * 0.10) if (houve_inadimplemento and not jg) else 0.0
        
        total_final_processo = base_multa + valor_multa + hon_523
        gerar_laudo_excel(processo, teve_transito, jg, df_danos, df_custas, (subtotal_princ+subtotal_juros), (hon_exigivel_princ+hon_exigivel_juros), valor_multa, hon_523, total_final_processo, arquivo_saida, houve_inadimplemento, termo_juros_raw, [], base_hon, prop_hon, prop_custas, hon_perc, hon_fixo)

    else:
        df_danos['Valor Atualizado'] = df_danos['Valor Histórico'] 
        df_custas['Exigível'] = df_custas['Valor Histórico'] if not jg else 0.0
        
        historico_cg = []; saldo_principal = 0.0; saldo_juros = 0.0
        data_corte = df_deducoes.iloc[0]['Data bloqueio/deposito']
        
        for idx, row in df_danos.iterrows():
            data_cm = row['Data Desembolso']
            if pd.isna(data_cm) or data_cm > data_corte: continue
            valor = float(row['Valor Histórico'])
            regra = str(row.get('Regra', '')).strip().upper()
            data_juros = data_citacao if 'CITA' in termo_juros_raw else data_evento if 'EVENTO' in termo_juros_raw else data_cm
                
            if is_fazenda: f_cm, f_jur = calc_fazenda_publica(df_bcb, data_cm, data_juros, data_corte)
            elif regra == 'R1': f_cm, f_jur = calc_tjmg_juros(tabela_tjmg, data_cm, data_juros, data_corte)
            elif regra == 'R4': f_cm, f_jur = calc_tjmg_leinova(tabela_tjmg, df_bcb, data_cm, data_juros, data_corte)
            elif regra == 'R6': f_cm, f_jur = calc_leinova_pura(df_bcb, data_cm, data_juros, data_corte)
            else: f_cm, f_jur = calc_selic_pura(df_bcb, data_cm, data_juros, data_corte)
            
            val_princ = valor * f_cm
            saldo_principal += val_princ; saldo_juros += val_princ * f_jur
            
        for idx, row in df_custas.iterrows():
            data_cm_c = row['Data Desembolso']
            if pd.isna(data_cm_c) or data_cm_c > data_corte or jg: continue
            f_cm, f_jur = calc_fazenda_publica(df_bcb, data_cm_c, data_transito_c if teve_transito else pd.NaT, data_corte) if is_fazenda else calc_leinova_pura(df_bcb, data_cm_c, data_transito_c if teve_transito else pd.NaT, data_corte) 
            val_princ = (row['Valor Histórico'] * f_cm) * prop_custas
            saldo_principal += val_princ; saldo_juros += val_princ * f_jur
            
        historico_cg.append((data_corte, "Subtotal (Principal + Juros)", saldo_principal, saldo_juros, 0.0, saldo_principal+saldo_juros))
        
        hon_suc_princ, hon_suc_jur = 0.0, 0.0
        if base_hon == 'VALOR DA CAUSA' and valor_causa > 0 and pd.notna(data_propositura) and not jg:
            f_cm_hon, _ = calc_fazenda_publica(df_bcb, data_propositura, pd.NaT, data_corte) if is_fazenda else calc_tjmg_leinova(tabela_tjmg, df_bcb, data_propositura, pd.NaT, data_corte)
            hon_suc_princ = ((valor_causa * f_cm_hon) * hon_perc) * prop_hon
        elif hon_fixo > 0 and pd.notna(data_sentenca) and not jg:
            f_cm, f_jur = calc_fazenda_publica(df_bcb, data_sentenca, data_transito_c if teve_transito else pd.NaT, data_corte) if is_fazenda else calc_leinova_pura(df_bcb, data_sentenca, data_transito_c if teve_transito else pd.NaT, data_corte)
            hon_suc_princ = (hon_fixo * f_cm) * prop_hon
            hon_suc_jur = hon_suc_princ * f_jur
        elif hon_fixo > 0 and not jg: hon_suc_princ = hon_fixo * prop_hon
        elif not jg: hon_suc_princ = ((saldo_principal + saldo_juros) * hon_perc) * prop_hon
            
        if hon_suc_princ > 0 or hon_suc_jur > 0:
            saldo_principal += hon_suc_princ; saldo_juros += hon_suc_jur
            historico_cg.append((data_corte, f"(+) Honorários Suc. ({f'{hon_perc*100:.0f}%' if hon_fixo==0 else 'Fixos'}{f' rateados a {prop_hon*100:.0f}%' if prop_hon<1.0 else ''})", hon_suc_princ, hon_suc_jur, 0.0, saldo_principal+saldo_juros))
            
        if houve_inadimplemento:
            base_multa = saldo_principal + saldo_juros
            multa_523 = base_multa * 0.10
            saldo_principal += multa_523
            historico_cg.append((data_corte, "(+) Multa Art. 523 CPC (10%)", multa_523, 0.0, 0.0, saldo_principal+saldo_juros))
            if not jg:
                hon_523 = base_multa * 0.10
                saldo_principal += hon_523
                historico_cg.append((data_corte, "(+) Honorários Art. 523 CPC (10%)", hon_523, 0.0, 0.0, saldo_principal+saldo_juros))
                
        historico_cg.append((data_corte, "DÍVIDA CONSOLIDADA (Pré-Bloqueio)", saldo_principal, saldo_juros, 0.0, saldo_principal+saldo_juros))
        
        ultima_data = data_corte
        for _, row in df_deducoes.iterrows():
            data_ded = pd.to_datetime(row['Data bloqueio/deposito'])
            valor_ded = float(row['Valor'])
            id_deposito = str(row.get('ID / Folha', row.get('ID', ''))).strip()
            str_id = f" (ID: {id_deposito})" if id_deposito and id_deposito.lower() != 'nan' else ""
            
            despesas_tardias = df_danos[(df_danos['Data Desembolso'] > ultima_data) & (df_danos['Data Desembolso'] <= data_ded)]
            for _, nd in despesas_tardias.iterrows():
                d_dano = nd['Data Desembolso']
                f_cm, f_jur = calc_fazenda_publica(df_bcb, ultima_data, ultima_data, d_dano) if is_fazenda else calc_leinova_pura(df_bcb, ultima_data, ultima_data, d_dano) 
                saldo_principal = saldo_principal * f_cm; saldo_juros += saldo_principal * f_jur
                saldo_principal += float(nd['Valor Histórico'])
                historico_cg.append((d_dano, f"(+) Nova Inclusão: {nd['Descrição']}", saldo_principal, saldo_juros, 0.0, saldo_principal+saldo_juros))
                ultima_data = d_dano

            if data_ded > ultima_data:
                f_cm, f_jur = calc_fazenda_publica(df_bcb, ultima_data, ultima_data, data_ded) if is_fazenda else calc_leinova_pura(df_bcb, ultima_data, ultima_data, data_ded) 
                saldo_principal = saldo_principal * f_cm; saldo_juros += saldo_principal * f_jur
                historico_cg.append((data_ded, "Atualização do Saldo", saldo_principal, saldo_juros, 0.0, saldo_principal+saldo_juros))
                ultima_data = data_ded
            
            abate_juros = min(valor_ded, saldo_juros)
            saldo_juros -= abate_juros
            abate_princ = valor_ded - abate_juros
            
            if abate_princ > saldo_principal:
                excesso = abate_princ - saldo_principal
                saldo_principal = 0.0
                historico_cg.append((data_ded, f"(-) Bloqueio/Depósito{str_id}", 0.0, saldo_juros, valor_ded, 0.0))
                historico_cg.append((data_ded, "(!) EXCESSO / RESTITUIR", 0.0, 0.0, excesso, 0.0))
            else:
                saldo_principal -= abate_princ
                historico_cg.append((data_ded, f"(-) Bloqueio/Depósito{str_id}", saldo_principal, saldo_juros, valor_ded, saldo_principal+saldo_juros))
        
        despesas_finais = df_danos[df_danos['Data Desembolso'] > ultima_data]
        for _, nd in despesas_finais.iterrows():
            d_dano = nd['Data Desembolso']
            f_cm, f_jur = calc_fazenda_publica(df_bcb, ultima_data, ultima_data, d_dano) if is_fazenda else calc_leinova_pura(df_bcb, ultima_data, ultima_data, d_dano) 
            saldo_principal = saldo_principal * f_cm; saldo_juros += saldo_principal * f_jur
            saldo_principal += float(nd['Valor Histórico'])
            historico_cg.append((d_dano, f"(+) Nova Inclusão: {nd['Descrição']}", saldo_principal, saldo_juros, 0.0, saldo_principal+saldo_juros))
            ultima_data = d_dano

        if ultima_data < data_calculo:
            f_cm, f_jur = calc_fazenda_publica(df_bcb, ultima_data, ultima_data, data_calculo) if is_fazenda else calc_leinova_pura(df_bcb, ultima_data, ultima_data, data_calculo)
            saldo_principal = saldo_principal * f_cm; saldo_juros += saldo_principal * f_jur
            historico_cg.append((data_calculo, "Atualização Final (Hoje)", saldo_principal, saldo_juros, 0.0, saldo_principal+saldo_juros))
            
        total_final_processo = saldo_principal + saldo_juros
        gerar_laudo_excel(processo, teve_transito, jg, df_danos, df_custas, 0, 0, 0, 0, total_final_processo, arquivo_saida, houve_inadimplemento, termo_juros_raw, historico_cg, base_hon, prop_hon, prop_custas, hon_perc, hon_fixo)

    # GERA EXITO APENAS SE FOR RÉU/DEFESA
    if 'AUTOR' not in atuacao and 'Risco Atual' in df_danos.columns and df_danos['Risco Atual'].sum() > 0:
        gerar_relatorio_exito_cliente(processo, df_danos[df_danos['Risco Atual'] > 0], arquivo_saida)

# --- 6. GERAÇÕES DE ARQUIVOS (LAUDO E ÊXITO) ---
def gerar_laudo_excel(processo, teve_transito, jg, df_danos, df_custas, subtotal, hon, multa, hon_523, total, arquivo_saida, houve_inadimplemento, termo_juros_raw, historico, base_hon, prop_hon, prop_custas, hon_perc, hon_fixo):
    wb = Workbook()
    ws = wb.active
    ws.title = "Laudo de Liquidação"

    f_titulo = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    f_negrito = Font(name="Arial", size=11, bold=True)
    f_normal = Font(name="Arial", size=11)
    fundo_escuro = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    fundo_cinza = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    borda = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    moeda = 'R$ #,##0.00'

    ws.merge_cells('A1:H1'); ws['A1'] = f"LAUDO DE CÁLCULO JUDICIAL - NASH SYSTEM"
    ws['A1'].font = f_titulo; ws['A1'].fill = fundo_escuro; ws['A1'].alignment = Alignment(horizontal="center")
    ws.merge_cells('A2:H2'); ws['A2'] = f"Gerado pelo Nash System v{__version__} em {pd.Timestamp.today().strftime('%d/%m/%Y às %H:%M')}"
    ws['A2'].alignment = Alignment(horizontal="right"); ws['A2'].font = Font(name="Arial", size=9, italic=True)

    ws['A4'] = "Processo:"; ws['B4'] = processo
    ws['A5'] = "Termo Juros (Danos):"; ws['B5'] = termo_juros_raw.title()
    ws['A6'] = "Justiça Gratuita:"; ws['B6'] = "DEFERIDA (Custas Inexigíveis)" if jg else "NÃO REQUERIDA / INDEFERIDA"
    ws['D4'] = "Base Honorários:"; ws['E4'] = base_hon.title()
    
    linha_cab = 5
    if prop_hon < 1.0: ws[f'D{linha_cab}'] = "Proporção Hon.:"; ws[f'E{linha_cab}'] = f"{prop_hon*100:.2f}%"; linha_cab += 1
    if prop_custas < 1.0: ws[f'D{linha_cab}'] = "Proporção Custas:"; ws[f'E{linha_cab}'] = f"{prop_custas*100:.2f}%"; linha_cab += 1
    
    ws[f'D{linha_cab}'] = "Perc. Honorários:"
    if hon_fixo > 0: ws[f'E{linha_cab}'] = f"Fixo R$ {hon_fixo:,.2f}".replace(',','_').replace('.',',').replace('_','.')
    else: ws[f'E{linha_cab}'] = f"{hon_perc*100:.0f}%"

    for r in range(4, 7): ws.cell(row=r, column=1).font = f_negrito; ws.cell(row=r, column=4).font = f_negrito

    linha = 8
    ws.merge_cells(f'A{linha}:H{linha}'); ws[f'A{linha}'] = "1. DANOS MATERIAIS E VALORES PRINCIPAIS"
    ws[f'A{linha}'].font = f_negrito; ws[f'A{linha}'].fill = fundo_cinza; linha += 1

    cabs = ['ID / Folha', 'Descrição', 'Data', 'Valor Hist.', 'Fator Corr.', 'Juros (%)', 'Regra', 'Valor Atualizado']
    for i, t in enumerate(cabs, 1): 
        ws.cell(row=linha, column=i, value=t).font = f_negrito
        ws.cell(row=linha, column=i).border = borda
        ws.cell(row=linha, column=i).alignment = Alignment(horizontal="center", vertical="center")
    linha += 1

    for _, r in df_danos.iterrows():
        exibe_data = r['Data Desembolso'].strftime('%d/%m/%Y') if float(r['Valor Histórico']) > 0 else "-" 
        ws.cell(row=linha, column=1, value=r['ID / Folha']).border = borda
        ws.cell(row=linha, column=2, value=r['Descrição']).border = borda
        ws.cell(row=linha, column=3, value=exibe_data).border = borda
        ws.cell(row=linha, column=4, value=r['Valor Histórico']).number_format = moeda; ws.cell(row=linha, column=4).border = borda
        ws.cell(row=linha, column=5, value=r.get('Fator CM', 1.0)).number_format = '0.0000000'; ws.cell(row=linha, column=5).border = borda
        ws.cell(row=linha, column=6, value=r.get('Fator Juros', 0.0)).number_format = '0.00%'; ws.cell(row=linha, column=6).border = borda
        ws.cell(row=linha, column=7, value=r.get('Desc_Regra', '')).border = borda
        val_display = r['Valor Histórico'] if historico else r['Valor Atualizado']
        ws.cell(row=linha, column=8, value=val_display).number_format = moeda; ws.cell(row=linha, column=8).border = borda
        linha += 1

    linha += 1
    ws.merge_cells(f'A{linha}:H{linha}'); ws[f'A{linha}'] = "2. CUSTAS E DESPESAS PROCESSUAIS"
    ws[f'A{linha}'].font = f_negrito; ws[f'A{linha}'].fill = fundo_cinza; linha += 1
    
    cabs_custas = ['ID / Folha', 'Descrição', 'Data', 'Valor Hist.', 'Fator Corr.', 'Juros (%)', 'Regra', 'Atualizado Exigível']
    for i, t in enumerate(cabs_custas, 1): 
        ws.cell(row=linha, column=i, value=t).font = f_negrito
        ws.cell(row=linha, column=i).border = borda
        ws.cell(row=linha, column=i).alignment = Alignment(horizontal="center", vertical="center")
    linha += 1
    
    for _, r in df_custas.iterrows():
        ws.cell(row=linha, column=1, value=r['ID / Folha']).border = borda
        ws.cell(row=linha, column=2, value=r['Descrição']).border = borda
        ws.cell(row=linha, column=3, value=r['Data Desembolso'].strftime('%d/%m/%Y')).border = borda
        ws.cell(row=linha, column=4, value=r['Valor Histórico']).number_format = moeda; ws.cell(row=linha, column=4).border = borda
        ws.cell(row=linha, column=5, value=r.get('Fator CM', 1.0)).number_format = '0.0000000'; ws.cell(row=linha, column=5).border = borda
        ws.cell(row=linha, column=6, value=r.get('Fator Juros', 0.0)).number_format = '0.00%'; ws.cell(row=linha, column=6).border = borda
        ws.cell(row=linha, column=7, value=r.get('Desc_Regra', '')).border = borda
        val_display_c = r['Valor Histórico'] if (historico and not jg) else r.get('Exigível', 0.0)
        ws.cell(row=linha, column=8, value=0.0 if jg else val_display_c).number_format = moeda; ws.cell(row=linha, column=8).border = borda
        linha += 1

    linha += 2

    if historico:
        ws.merge_cells(f'A{linha}:H{linha}'); ws[f'A{linha}'] = "3. EVOLUÇÃO DA DÍVIDA (AMORTIZAÇÃO CONTA GRÁFICA - Art. 354 CC)"
        ws[f'A{linha}'].font = f_titulo; ws[f'A{linha}'].fill = fundo_escuro; ws[f'A{linha}'].alignment = Alignment(horizontal="center"); linha += 1
        
        ws.merge_cells(f'A{linha}:B{linha}'); ws.merge_cells(f'C{linha}:D{linha}')
        cabs_cg_titles = ['Data', 'Evento', 'Principal Corrigido', 'Juros Acumulados', 'Valor Pago/Bloqueio', 'Saldo Devedor Total']
        
        ws.cell(row=linha, column=1, value='Data').font = f_negrito; ws.cell(row=linha, column=1).border = borda; ws.cell(row=linha, column=1).fill = fundo_cinza; ws.cell(row=linha, column=1).alignment = Alignment(vertical="center", horizontal="center")
        ws.cell(row=linha, column=2).border = borda; ws.cell(row=linha, column=2).fill = fundo_cinza
        ws.cell(row=linha, column=3, value='Evento').font = f_negrito; ws.cell(row=linha, column=3).border = borda; ws.cell(row=linha, column=3).fill = fundo_cinza; ws.cell(row=linha, column=3).alignment = Alignment(vertical="center", horizontal="center")
        ws.cell(row=linha, column=4).border = borda; ws.cell(row=linha, column=4).fill = fundo_cinza
        
        for i, t in enumerate(cabs_cg_titles[2:], 5): 
            ws.cell(row=linha, column=i, value=t).font = f_negrito
            ws.cell(row=linha, column=i).border = borda
            ws.cell(row=linha, column=i).fill = fundo_cinza
            ws.cell(row=linha, column=i).alignment = Alignment(horizontal="center", vertical="center")
        linha += 1
        
        for data, evento, princ, jur, pago, saldo in historico:
            ws.merge_cells(f'A{linha}:B{linha}'); ws.merge_cells(f'C{linha}:D{linha}')
            ws.cell(row=linha, column=1, value=data.strftime('%d/%m/%Y')).border = borda; ws.cell(row=linha, column=2).border = borda
            ws.cell(row=linha, column=3, value=evento).border = borda; ws.cell(row=linha, column=4).border = borda
            
            if "Restituir" in str(evento).title(): ws.cell(row=linha, column=3).font = Font(name="Arial", size=11, bold=True, color="FF0000")
            elif "Bloqueio" in str(evento) or "Depósito" in str(evento): ws.cell(row=linha, column=3).font = f_negrito
            else: ws.cell(row=linha, column=3).font = f_normal
                
            ws.cell(row=linha, column=5, value=princ).number_format = moeda; ws.cell(row=linha, column=5).border = borda
            ws.cell(row=linha, column=6, value=jur).number_format = moeda; ws.cell(row=linha, column=6).border = borda
            ws.cell(row=linha, column=7, value=pago if pago > 0 else "-").number_format = moeda if pago > 0 else 'General'; ws.cell(row=linha, column=7).border = borda
            ws.cell(row=linha, column=8, value=saldo).number_format = moeda; ws.cell(row=linha, column=8).border = borda; ws.cell(row=linha, column=8).font = f_negrito
            linha += 1
            
        linha += 1; ws.merge_cells(f'A{linha}:G{linha}'); ws.cell(row=linha, column=1, value="SALDO DEVEDOR FINAL DA EXECUÇÃO:").alignment = Alignment(horizontal="right")
        ws.cell(row=linha, column=1).font = f_titulo; ws.cell(row=linha, column=1).fill = fundo_escuro
        ws.cell(row=linha, column=8, value=(historico[-1][5] if historico else 0.0)).number_format = moeda
        ws.cell(row=linha, column=8).font = f_titulo; ws.cell(row=linha, column=8).fill = fundo_escuro
        for i in range(1, 9): ws.cell(row=linha, column=i).border = borda
            
    else:
        ws.merge_cells(f'A{linha}:H{linha}'); ws[f'A{linha}'] = "3. RESUMO DA LIQUIDAÇÃO"
        ws[f'A{linha}'].font = f_titulo; ws[f'A{linha}'].fill = fundo_escuro; ws[f'A{linha}'].alignment = Alignment(horizontal="center"); linha += 1

        def add_total(desc, val, negrito=False, dest=False):
            nonlocal linha; ws.merge_cells(f'A{linha}:G{linha}')
            ws.cell(row=linha, column=1, value=desc).alignment = Alignment(horizontal="right"); ws.cell(row=linha, column=8, value=val).number_format = moeda
            fonte = f_negrito if negrito else f_normal
            ws.cell(row=linha, column=1).font = fonte; ws.cell(row=linha, column=8).font = fonte
            if dest: ws.cell(row=linha, column=1).fill = fundo_cinza; ws.cell(row=linha, column=8).fill = fundo_cinza
            for i in range(1, 9): ws.cell(row=linha, column=i).border = borda
            linha += 1

        add_total("SUBTOTAL (Principal + Custas Atualizados):", subtotal, True)
        add_total(f"Honorários Acumulados ({f'{hon_perc*100:.0f}%' if hon_fixo==0 else 'Fixos'}){f' (rateado a {prop_hon*100:.0f}%)' if prop_hon<1.0 else ''}:", hon)
        if houve_inadimplemento:
            add_total("Multa Art. 523 CPC (10%):", multa); add_total("Honorários Art. 523 CPC (10%):", hon_523)
        add_total("TOTAL GERAL DEVIDO:", total, True, True)

    larguras_minimas = {'A': 16, 'B': 28, 'C': 16, 'D': 16, 'E': 14, 'F': 12, 'G': 18, 'H': 20}
    for letra_col, larg_min in larguras_minimas.items():
        ws.column_dimensions[letra_col].width = larg_min
    
    ws.print_area = f'A1:H{linha}'; ws.page_setup.fitToWidth = 1; ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    
    nome_saida = str(arquivo_saida)
    if nome_saida.endswith(('.ods', '.xls')): nome_saida = nome_saida.rsplit('.', 1)[0] + '.xlsx'
    path_final = Path(nome_saida)
    if path_final.exists():
        pasta = path_final.parent; nome_base = path_final.stem; ext = path_final.suffix; contador = 1
        while True:
            novo_caminho = pasta / f"{nome_base} ({contador}){ext}"
            if not novo_caminho.exists(): path_final = novo_caminho; break
            contador += 1
            
    wb.save(str(path_final))
    converter_para_pdf(path_final)

def gerar_relatorio_exito_cliente(processo, df_exito, arquivo_saida_base):
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório de Êxito"
    
    f_titulo = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    f_negrito = Font(name="Arial", size=11, bold=True)
    f_normal = Font(name="Arial", size=11)
    f_sucesso = Font(name="Arial", size=12, bold=True, color="107C41") 
    
    fundo_escuro = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid") 
    fundo_cinza = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    fundo_verde = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    borda = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    moeda = 'R$ #,##0.00'

    ws.merge_cells('A1:D1'); ws['A1'] = "DEMONSTRATIVO DE PROVEITO ECONÔMICO (ÊXITO)"
    ws['A1'].font = f_titulo; ws['A1'].fill = fundo_escuro; ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    
    ws.merge_cells('A2:D2'); ws['A2'] = f"Processo: {processo}"
    ws['A2'].alignment = Alignment(horizontal="center"); ws['A2'].font = Font(name="Arial", size=10, italic=True)

    ws.append([]) 
    
    cabs = ['Descrição da Verba (Pedidos)', 'Risco Atualizado (O que o autor pediu)', 'Condenação Atualizada (O que o autor levou)', 'Economia Gerada (Proveito do Cliente)']
    for col, texto in enumerate(cabs, 1):
        ws.cell(row=4, column=col, value=texto).font = f_negrito
        ws.cell(row=4, column=col).fill = fundo_cinza
        ws.cell(row=4, column=col).border = borda
        ws.cell(row=4, column=col).alignment = Alignment(horizontal="center", vertical="center")
        
    linha = 5
    soma_risco = soma_condenacao = soma_proveito = 0.0
    
    for _, r in df_exito.iterrows():
        ws.cell(row=linha, column=1, value=r['Descrição']).border = borda
        ws.cell(row=linha, column=2, value=r['Risco Atual']).number_format = moeda; ws.cell(row=linha, column=2).border = borda
        ws.cell(row=linha, column=3, value=r['Valor Atualizado']).number_format = moeda; ws.cell(row=linha, column=3).border = borda
        ws.cell(row=linha, column=4, value=r['Proveito']).number_format = moeda; ws.cell(row=linha, column=4).border = borda
        ws.cell(row=linha, column=4).font = f_sucesso; ws.cell(row=linha, column=4).fill = fundo_verde
        
        soma_risco += r['Risco Atual']; soma_condenacao += r['Valor Atualizado']
        soma_proveito += r['Proveito']
        linha += 1
        
    ws.append([])
    linha += 1
    
    ws.cell(row=linha, column=1, value="TOTAIS GERAIS:").font = Font(name="Arial", size=12, bold=True)
    ws.cell(row=linha, column=1).alignment = Alignment(horizontal="right")
    ws.cell(row=linha, column=2, value=soma_risco).number_format = moeda; ws.cell(row=linha, column=2).font = f_negrito
    ws.cell(row=linha, column=3, value=soma_condenacao).number_format = moeda; ws.cell(row=linha, column=3).font = f_negrito
    
    ws.cell(row=linha, column=4, value=soma_proveito).number_format = moeda; ws.cell(row=linha, column=4).font = f_sucesso
    ws.cell(row=linha, column=4).fill = fundo_verde
    
    for c in range(1, 5): ws.cell(row=linha, column=c).border = borda

    ws.column_dimensions['A'].width = 35; ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 35; ws.column_dimensions['D'].width = 35
    
    ws.print_area = f'A1:D{linha}'; ws.page_setup.fitToWidth = 1; ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    
    nome_base = arquivo_saida_base.name.replace('Laudo_', 'Relatorio_Exito_Detalhado_')
    if 'Relatorio_Exito_' not in nome_base: nome_base = f"Relatorio_Exito_Detalhado_{arquivo_saida_base.name}"
    
    path_final = arquivo_saida_base.parent / nome_base
    if path_final.exists():
        pasta, nome_raiz, ext, contador = path_final.parent, path_final.stem, path_final.suffix, 1
        while True:
            novo_caminho = pasta / f"{nome_raiz} ({contador}){ext}"
            if not novo_caminho.exists(): path_final = novo_caminho; break
            contador += 1
            
    wb.save(str(path_final))
    logging.info(f"Relatório de Êxito Cliente gerado com sucesso: {path_final.name}")
    converter_para_pdf(path_final)

class NashGUI:
    def __init__(self, root):
        self.root = root
        self.root.title(f"Nash System - Liquidação (v{__version__})")
        self.root.geometry("640x580")
        self.root.configure(padx=20, pady=20)
        icone_path = PASTA_TEMP / "dr_nash.ico"
        if icone_path.exists():
            try: self.root.iconbitmap(str(icone_path))
            except: pass

        tk.Label(root, text="NASH SYSTEM", font=("Arial", 16, "bold")).pack(pady=(0, 5))
        tk.Label(root, text="Assistente de Cálculos Judiciais", font=("Arial", 10, "italic")).pack(pady=(0, 15))

        frame_regras = tk.LabelFrame(root, text=" 📖 Dicionário de Regras Matemáticas ", font=("Arial", 10, "bold"), padx=10, pady=8)
        frame_regras.pack(fill="x", pady=5)
        regras = [
            ("R1", "TJMG + Juros de 1% a.m."),
            ("R2", "Taxa Selic (critério único)"),
            ("R3", "TJMG + Juros 1% a.m. até 08/24; após, Selic"),
            ("R4", "TJMG + Juros 1% a.m. até 08/24; após, Lei 14.905/24"),
            ("R5", "Selic até 08/24; após, Lei 14.905/24 (Tema 1.368 STJ)"),
            ("R6", "Lei 14.905/24: IPCA + Taxa Legal"),
            ("Hon.", "Base na Condenação ou Valor da Causa (STJ: s/ juros)"),
            ("Prop.", "Reembolso proporcional de Custas/Hon. (ex: 70%)"),
            ("Exito", "Gera Relat. de Economia aut. se 'Valor Pedido' em Danos.")
        ]
        for regra, desc in regras:
            linha_ui = tk.Frame(frame_regras)
            linha_ui.pack(anchor="w", pady=1)
            tk.Label(linha_ui, text=f"{regra}: ", font=("Arial", 9, "bold")).pack(side="left")
            tk.Label(linha_ui, text=desc, font=("Arial", 9)).pack(side="left")

        self.lbl_status = tk.Label(root, text="Aguardando arquivo...", font=("Arial", 10), fg="gray")
        self.lbl_status.pack(pady=10)
        self.btn_processar = tk.Button(root, text="📂 Selecionar Planilha e Calcular", font=("Arial", 11, "bold"), bg="#2F5597", fg="white", padx=15, pady=8, command=self.iniciar_processo)
        self.btn_processar.pack()

    def iniciar_processo(self):
        caminho = filedialog.askopenfilename(title="Selecione a Planilha", filetypes=[("Planilhas", "*.xlsx *.xls *.ods")])
        if not caminho: return
        caminho_entrada = Path(caminho)
        arquivo_saida = caminho_entrada.parent / f"Laudo_{caminho_entrada.name}"
        self.lbl_status.config(text="Processando cálculo e gerando PDFs...", fg="blue")
        self.btn_processar.config(state="disabled")
        self.root.update()
        threading.Thread(target=self.processar_em_background, args=(caminho_entrada, arquivo_saida)).start()

    def processar_em_background(self, caminho_entrada, arquivo_saida):
        logging.info(f"Iniciando cálculo para: {caminho_entrada.name}")
        try:
            executar_nash(caminho_entrada, arquivo_saida)
            def sucesso():
                self.lbl_status.config(text=f"Cálculos e PDFs gerados com sucesso.", fg="green")
                self.btn_processar.config(state="normal")
                messagebox.showinfo("Sucesso", "O Laudo de Liquidação e Êxito foram processados e exportados para PDF!")
            self.root.after(0, sucesso)
        except Exception as e:
            tb_str = traceback.format_exc()
            logging.error(f"Erro ao processar {caminho_entrada.name}:\n{tb_str}")
            erro_msg = str(e)
            def erro():
                self.lbl_status.config(text="Erro crítico (Veja o log).", fg="red")
                self.btn_processar.config(state="normal")
                messagebox.showerror("Atenção - Erro no Processamento", f"O cálculo falhou:\n\n{erro_msg}") 
            self.root.after(0, erro)

if __name__ == "__main__":
    app = tk.Tk()
    gui = NashGUI(app)
    app.mainloop()