import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --- 1. CONFIGURAÇÃO DAS PASTAS ---
PASTA_BASE = Path(__file__).parent
ARQUIVO_TJMG = PASTA_BASE / 'Tabelas_Oficiais' / 'tabela_tjmg.xlsx'
ARQUIVO_ENTRADA = PASTA_BASE / 'Processos_Entrada' / 'planilha_teste.xlsx'
ARQUIVO_SAIDA = PASTA_BASE / 'Processos_Calculados' / 'planilha_teste_calculada.xlsx'

# --- 2. MOTOR DO TJMG (Reutilizando a inteligência que já criamos) ---
def carregar_tabela_tjmg():
    df = pd.read_excel(ARQUIVO_TJMG, sheet_name='Plan1', skiprows=8, names=['ANO', 'MÊS', 'ÍNDICE'])
    df = df.dropna(subset=['MÊS', 'ÍNDICE'])
    df['MÊS'] = df['MÊS'].astype(str).str.strip()
    meses = {'Janeiro': '01', 'Fevereiro': '02', 'Março': '03', 'Abril': '04', 'Maio': '05', 'Junho': '06', 'Julho': '07', 'Agosto': '08', 'Setembro': '09', 'Outubro': '10', 'Novembro': '11', 'Dezembro': '12'}
    df['MÊS_NUM'] = df['MÊS'].map(meses)
    df = df.dropna(subset=['MÊS_NUM'])
    df['DATA_REF'] = pd.to_datetime(df['ANO'].astype(int).astype(str) + '-' + df['MÊS_NUM'] + '-01')
    return df[['DATA_REF', 'ÍNDICE']].set_index('DATA_REF')

def calcular_fator_tjmg(df_tjmg, data_nota, data_limite="2024-08-01"):
    # Converte para o dia 01 do mês para buscar na tabela do TJMG
    data_nota_mes = pd.to_datetime(f"{data_nota.year}-{data_nota.month:02d}-01")
    data_limite_mes = pd.to_datetime(data_limite)
    try:
        return df_tjmg.loc[data_nota_mes, 'ÍNDICE'] / df_tjmg.loc[data_limite_mes, 'ÍNDICE']
    except KeyError:
        return 1.0 # Se a data não for encontrada (ex: data futura), não corrige

# --- 3. PROCESSAMENTO DA PLANILHA ---
print(f"Lendo processos de: {ARQUIVO_ENTRADA.name}...")
try:
    df_tjmg = carregar_tabela_tjmg()
    df_processo = pd.read_excel(ARQUIVO_ENTRADA)
    
    # Criando a nova coluna de resultados
    df_processo['VALOR ATUALIZADO (AGO/24)'] = 0.0
    df_processo['OBSERVAÇÃO'] = ""

    # Iterando linha por linha como um contador faria
    for index, row in df_processo.iterrows():
        JG = str(row.get('JUSTIÇA GRATUITA', '')).strip().upper()
        
        if JG == 'SIM':
            df_processo.at[index, 'VALOR ATUALIZADO (AGO/24)'] = row['VALOR HISTÓRICO']
            df_processo.at[index, 'OBSERVAÇÃO'] = "Exigibilidade Suspensa (JG)"
            continue
            
        regra = str(row.get('REGRA', '')).strip()
        data_desembolso = pd.to_datetime(row['DATA DESEMBOLSO'], dayfirst=True)
        valor_historico = float(row['VALOR HISTÓRICO'])
        
        if regra == 'TJMG_ate_LeiNova':
            fator = calcular_fator_tjmg(df_tjmg, data_desembolso)
            valor_corrigido = valor_historico * fator
            df_processo.at[index, 'VALOR ATUALIZADO (AGO/24)'] = valor_corrigido
            df_processo.at[index, 'OBSERVAÇÃO'] = f"Fator TJMG: {fator:.6f}"
        else:
            df_processo.at[index, 'VALOR ATUALIZADO (AGO/24)'] = valor_historico
            df_processo.at[index, 'OBSERVAÇÃO'] = "Regra não processada ainda"

    # Salva o Excel cru
    df_processo.to_excel(ARQUIVO_SAIDA, index=False)
    print("Cálculos finalizados. Iniciando formatação visual...")

except Exception as e:
    print(f"Erro no processamento: {e}")
    exit()

# --- 4. FORMATAÇÃO VISUAL OFICIAL (openpyxl) ---
wb = load_workbook(ARQUIVO_SAIDA)
ws = wb.active

# Estilos
fundo_azul = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
fonte_branca = Font(color="FFFFFF", bold=True)
borda_fina = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
formato_moeda = 'R$ #,##0.00'

# Pintando o Cabeçalho
for col in range(1, ws.max_column + 1):
    celula = ws.cell(row=1, column=col)
    celula.fill = fundo_azul
    celula.font = fonte_branca
    celula.alignment = Alignment(horizontal="center", vertical="center")
    celula.border = borda_fina

# Formatando as linhas de dados e ajustando larguras
for row in range(2, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        celula = ws.cell(row=row, column=col)
        celula.border = borda_fina
        
        # Se for coluna de Valor (coluna 5 e coluna 8)
        if col in [5, 8]:
            celula.number_format = formato_moeda

# Ajuste automático das larguras das colunas
for col in ws.columns:
    max_length = 0
    col_letter = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    ws.column_dimensions[col_letter].width = max_length + 2

wb.save(ARQUIVO_SAIDA)
print(f"✓ SUCESSO! A planilha formatada foi salva em: {ARQUIVO_SAIDA.name}")